
import telebot
import config
import random

from openpyxl import load_workbook
from telebot import types

bot = telebot.TeleBot(config.TOKEN1)


@bot.message_handler(commands=['start'])
def welcome(message):

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Японская 🇯🇵")
    item2 = types.KeyboardButton("Европейская 🌍")
    item3 = types.KeyboardButton("Русская 🇷🇺")
    markup.add(item1, item2, item3)

    bot.send_message(message.chat.id, "Привет, {0.first_name}, выбирай кухню которая по вкусу.".format(message.from_user, bot.get_me()),
                     parse_mode='html', reply_markup=markup)


@bot.message_handler(func=lambda message: True)
def menu(message):
    if message.chat.type == "private":
        if message.text == "Японская 🇯🇵":
            wb = load_workbook(filename='static/Kitchen.xlsx')
            sheet_ranges = wb['Япония']
            info = "Наш выбор пал на 🥁 : \n " + str(sheet_ranges['A' + str(random.randint(1, 12))].value)
            bot.send_message(message.chat.id, info)

        elif message.text == "Европейская 🌍":
            wb = load_workbook(filename='static/Kitchen.xlsx')
            sheet_ranges = wb['Европа']
            info = "Наш выбор пал на 🥁 : \n " + str(sheet_ranges['A' + str(random.randint(1, 12))].value)
            bot.send_message(message.chat.id, info)

        elif message.text == "Русская 🇷🇺":
            wb = load_workbook(filename='static/Kitchen.xlsx')
            sheet_ranges = wb['Россия']
            info = "Наш выбор пал на 🥁 : \n " + str(sheet_ranges['A' + str(random.randint(1, 12))].value)
            bot.send_message(message.chat.id, info)

        else:
            bot.send_message(message.chat.id, 'Выберите кухню.')


bot.polling(none_stop=True)
